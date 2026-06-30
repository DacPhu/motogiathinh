"""Reports — dashboard PDF, 7-day PDF, full data Excel.

Uses reportlab for PDFs and openpyxl for Excel.
"""

from __future__ import annotations

import os
from collections import defaultdict
from datetime import datetime, timedelta, timezone
from decimal import Decimal
from io import BytesIO

from fastapi import APIRouter, HTTPException, status
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from reportlab.lib import colors as rl_colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import cm
from reportlab.lib.styles import ParagraphStyle
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from sqlalchemy import select

from app.core.cache import CacheKeys, cache
from app.dependencies import CurrentUser, DB
from app.models.class_model import Class, ClassEnrollment
from app.models.enums import RoleName
from app.models.payment import Payment
from app.models.student import Student
from app.models.user import User
from app.routers import reports_xlsx as rx

router = APIRouter(prefix="/reports", tags=["reports"])

_CYAN   = rl_colors.HexColor("#1a7ba4")
_HEADER = rl_colors.HexColor("#0d3b52")
_ALT    = rl_colors.HexColor("#f0f7fb")
_WHITE  = rl_colors.white
_GRAY   = rl_colors.HexColor("#888888")


# ─── Vietnamese-capable PDF font ──────────────────────────────────────────────
# reportlab's built-in Helvetica is Latin-1 only — Vietnamese diacritics (ố, ữ,
# ạ…) render as boxes/blanks. We register DejaVu Sans (free, conventional sans-
# serif, full Vietnamese coverage incl. the đồng sign), BUNDLED in the repo at
# app/assets/fonts/ so it's always present at a fixed path — no dependence on an
# apt package landing fonts where we expect. System paths are extra fallbacks for
# local dev. If NOTHING usable is found we raise: a broken-glyph PDF is worse than
# a clear error, and the bundled font means that branch should be unreachable.
_FONT_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), "assets", "fonts")


def _register_vn_font() -> tuple[str, str]:
    candidates = [
        # (regular, bold) — first existing pair wins. Bundled font is first:
        # resolved relative to this module, so it's guaranteed to exist.
        (os.path.join(_FONT_DIR, "DejaVuSans.ttf"),
         os.path.join(_FONT_DIR, "DejaVuSans-Bold.ttf")),
        # System fonts (local dev fallbacks)
        ("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
         "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf"),
        ("/usr/share/fonts/truetype/noto/NotoSans-Regular.ttf",
         "/usr/share/fonts/truetype/noto/NotoSans-Bold.ttf"),
        ("C:/Windows/Fonts/arial.ttf", "C:/Windows/Fonts/arialbd.ttf"),
        ("/Library/Fonts/Arial.ttf", "/Library/Fonts/Arial Bold.ttf"),
    ]
    for reg, bold in candidates:
        if os.path.exists(reg) and os.path.exists(bold):
            try:
                pdfmetrics.registerFont(TTFont("VN", reg))
                pdfmetrics.registerFont(TTFont("VN-Bold", bold))
                return "VN", "VN-Bold"
            except Exception:
                continue
    raise RuntimeError(
        f"No Vietnamese-capable PDF font found. Expected bundled fonts at {_FONT_DIR} "
        "(DejaVuSans.ttf + DejaVuSans-Bold.ttf)."
    )


_FONT, _FONT_BOLD = _register_vn_font()


def _vn_date(dt) -> str:
    if dt is None:
        return ""
    if hasattr(dt, "strftime"):
        return dt.strftime("%d/%m/%Y")
    return str(dt)


def _money(v) -> str:
    if v is None:
        return ""
    return f"{int(v):,}".replace(",", ".")


def _pdf_resp(buf: BytesIO, name: str) -> StreamingResponse:
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/pdf",
                             headers={"Content-Disposition": f'attachment; filename="{name}"'})


def _xlsx_resp(buf: BytesIO, name: str) -> StreamingResponse:
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{name}"'},
    )


def _tbl_style(header_rows: int = 1) -> list:
    return [
        ("BACKGROUND",    (0, 0), (-1, header_rows - 1), _HEADER),
        ("TEXTCOLOR",     (0, 0), (-1, header_rows - 1), _WHITE),
        ("FONTNAME",      (0, 0), (-1, header_rows - 1), _FONT_BOLD),
        ("FONTSIZE",      (0, 0), (-1, header_rows - 1), 8),
        ("ROWBACKGROUNDS",(0, header_rows), (-1, -1), [_WHITE, _ALT]),
        ("FONTNAME",      (0, header_rows), (-1, -1), _FONT),
        ("FONTSIZE",      (0, header_rows), (-1, -1), 7.5),
        ("GRID",          (0, 0), (-1, -1), 0.25, rl_colors.HexColor("#cccccc")),
        ("TOPPADDING",    (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("LEFTPADDING",   (0, 0), (-1, -1), 4),
        ("RIGHTPADDING",  (0, 0), (-1, -1), 4),
    ]


def _enum_val(v) -> str:
    return str(v.value if hasattr(v, "value") else v)


# ─── Dashboard PDF ──────────────────────────────────────────────────────────
@router.get("/dashboard.pdf")
async def dashboard_pdf(current_user: CurrentUser, db: DB):
    if current_user.role not in (RoleName.admin, RoleName.staff):
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="staff_only")
    from app.models.branch import Branch

    now = datetime.now(timezone.utc)
    students = (await db.execute(select(Student).where(Student.deleted_at.is_(None)))).scalars().all()
    payments = (await db.execute(select(Payment).where(Payment.deleted_at.is_(None)))).scalars().all()
    branches = (await db.execute(select(Branch))).scalars().all()

    branch_map = {b.id: b.ten_chi_nhanh for b in branches}
    total_rev  = sum(p.so_tien for p in payments if p.so_tien > 0)

    by_status: dict[str, int] = {}
    for s in students:
        k = _enum_val(s.trang_thai)
        by_status[k] = by_status.get(k, 0) + 1

    by_branch: dict[str, dict] = {str(b.id): {"name": b.ten_chi_nhanh, "students": 0, "revenue": Decimal(0)} for b in branches}
    for s in students:
        k = str(s.branch_id)
        if k in by_branch:
            by_branch[k]["students"] += 1
    for p in payments:
        k = str(p.branch_id)
        if k in by_branch:
            by_branch[k]["revenue"] += p.so_tien

    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=1.5*cm, rightMargin=1.5*cm, topMargin=1.5*cm, bottomMargin=1.5*cm)
    T  = ParagraphStyle
    story = [
        Paragraph("TỔNG QUAN HỆ THỐNG", T("tit", fontSize=16, spaceAfter=4, textColor=_HEADER, fontName=_FONT_BOLD)),
        Paragraph(f"Moto Gia Thịnh · {_vn_date(now)}", T("sub", fontSize=9, textColor=_GRAY, spaceAfter=12, fontName=_FONT)),
    ]

    kpi = [["CHỈ SỐ", "GIÁ TRỊ"],
           ["Tổng học viên", str(len(students))],
           ["Tổng doanh thu", _money(total_rev) + " đ"],
           ["Số chi nhánh", str(len(branches))]] + \
          [[rx._STATUS_VN.get(k, k), str(v)] for k, v in by_status.items()]
    t = Table(kpi, colWidths=[9*cm, 6*cm])
    t.setStyle(TableStyle(_tbl_style()))
    story += [t, Spacer(1, 0.5*cm),
              Paragraph("Theo chi nhánh", T("h2", fontSize=11, spaceAfter=4, textColor=_HEADER, fontName=_FONT_BOLD))]

    brows = [["CHI NHÁNH", "HỌC VIÊN", "DOANH THU (đ)"]]
    brows += [[v["name"], str(v["students"]), _money(v["revenue"])] for v in by_branch.values()]
    t2 = Table(brows, colWidths=[8*cm, 4*cm, 6*cm])
    t2.setStyle(TableStyle(_tbl_style()))
    story.append(t2)

    doc.build(story)
    return _pdf_resp(buf, f"tongquan-{now.strftime('%Y%m%d')}.pdf")


# ─── 7-day PDF ──────────────────────────────────────────────────────────────
@router.get("/data.pdf")
async def data_pdf(current_user: CurrentUser, db: DB):
    if current_user.role not in (RoleName.admin, RoleName.staff):
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="staff_only")
    now   = datetime.now(timezone.utc)
    since = now - timedelta(days=7)

    students_7 = (await db.execute(
        select(Student).where(Student.deleted_at.is_(None), Student.created_at >= since).order_by(Student.created_at.desc())
    )).scalars().all()
    payments_7 = (await db.execute(
        select(Payment).where(Payment.deleted_at.is_(None), Payment.collected_at >= since).order_by(Payment.collected_at.desc())
    )).scalars().all()
    # Resolve payment → student code (the old str(uuid)[:8]+"…" was meaningless).
    _sid = {p.student_id for p in payments_7 if p.student_id}
    smap = {s.id: s for s in (await db.execute(
        select(Student).where(Student.id.in_(_sid))
    )).scalars().all()} if _sid else {}

    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), leftMargin=1.5*cm, rightMargin=1.5*cm, topMargin=1.5*cm, bottomMargin=1.5*cm)
    T = ParagraphStyle
    story = [
        Paragraph("BÁO CÁO 7 NGÀY", T("tit", fontSize=14, spaceAfter=4, textColor=_HEADER, fontName=_FONT_BOLD)),
        Paragraph(f"{_vn_date(since)} – {_vn_date(now)}", T("sub", fontSize=9, textColor=_GRAY, spaceAfter=8, fontName=_FONT)),
        Paragraph(f"Học viên đăng ký ({len(students_7)})", T("h2", fontSize=10, spaceAfter=4, textColor=_HEADER, fontName=_FONT_BOLD)),
    ]

    srows = [["MÃ HV", "HỌ TÊN", "SĐT", "BẰNG LÁI", "TRẠNG THÁI", "NGÀY ĐK"]]
    for s in students_7:
        srows.append([s.ma_hoc_vien, s.ten_hoc_vien, s.so_dien_thoai,
                      _enum_val(s.loai_bang_lai),
                      rx._STATUS_VN.get(_enum_val(s.trang_thai), _enum_val(s.trang_thai)),
                      _vn_date(s.created_at)])
    t = Table(srows, colWidths=[2.5*cm, 5*cm, 3*cm, 2*cm, 3*cm, 2.5*cm])
    t.setStyle(TableStyle(_tbl_style()))
    story += [t, Spacer(1, 0.4*cm)]

    total_p = sum(p.so_tien for p in payments_7 if p.so_tien > 0)
    story.append(Paragraph(f"Thanh toán ({len(payments_7)}) — Tổng: {_money(total_p)} đ",
                           T("h2", fontSize=10, spaceAfter=4, textColor=_HEADER, fontName=_FONT_BOLD)))
    prows = [["MÃ GD", "MÃ HV", "SỐ TIỀN (đ)", "PHƯƠNG THỨC", "NGÀY THU"]]
    for p in payments_7:
        prows.append([p.ma_giao_dich,
                      (smap[p.student_id].ma_hoc_vien if p.student_id in smap else "—"),
                      _money(p.so_tien),
                      rx._METHOD_VN.get(_enum_val(p.phuong_thuc), _enum_val(p.phuong_thuc)),
                      _vn_date(p.collected_at)])
    t2 = Table(prows, colWidths=[3.5*cm, 3*cm, 3.5*cm, 3*cm, 2.5*cm])
    t2.setStyle(TableStyle(_tbl_style()))
    story.append(t2)

    doc.build(story)
    return _pdf_resp(buf, f"baocao-7ngay-{now.strftime('%Y%m%d')}.pdf")


# ─── Full Excel (6 sheets) ──────────────────────────────────────────────────
@router.get("/data.xlsx")
async def data_xlsx(current_user: CurrentUser, db: DB):
    if current_user.role not in (RoleName.admin, RoleName.staff):
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="staff_only")
    from app.models.branch import Branch

    now_dt = datetime.now(timezone.utc)

    # ── Load all data ──────────────────────────────────────────────────────────
    students = (await db.execute(
        select(Student).where(Student.deleted_at.is_(None)).order_by(Student.created_at.desc())
    )).scalars().all()
    payments = (await db.execute(
        select(Payment).where(Payment.deleted_at.is_(None)).order_by(Payment.collected_at.desc())
    )).scalars().all()
    branches = (await db.execute(select(Branch))).scalars().all()
    classes  = (await db.execute(
        select(Class).where(Class.deleted_at.is_(None)).order_by(Class.ngay_khai_giang.desc())
    )).scalars().all()
    all_users = (await db.execute(select(User))).scalars().all()

    # ── Lookup maps ────────────────────────────────────────────────────────────
    branch_map  = {b.id: b.ten_chi_nhanh for b in branches}
    student_map = {s.id: s for s in students}
    user_map    = {u.id: u for u in all_users}
    ctvs        = [u for u in all_users
                   if u.role == RoleName.collaborator and u.is_active and not u.deleted_at]
    ctv_map     = {u.id: u for u in ctvs}

    # student → primary class label (for the Học viên sheet "Lớp học" column)
    class_by_id = {c.id: c for c in classes}
    enrollments = (await db.execute(
        select(ClassEnrollment).where(ClassEnrollment.deleted_at.is_(None))
    )).scalars().all()
    class_map: dict = {}
    for e in enrollments:
        c = class_by_id.get(e.class_id)
        if c and e.student_id not in class_map:
            class_map[e.student_id] = c.ma_lop or c.ten_lop or ""

    # ── Per-student positive payment totals ────────────────────────────────────
    pos_paid: dict = defaultdict(int)
    for p in payments:
        if p.so_tien and int(p.so_tien) > 0 and p.student_id:
            pos_paid[p.student_id] += int(p.so_tien)

    # ── Summary aggregates ─────────────────────────────────────────────────────
    status_counts: dict = defaultdict(int)
    for s in students:
        status_counts[_enum_val(s.trang_thai)] += 1

    total_paid        = sum(int(p.so_tien) for p in payments if p.so_tien and int(p.so_tien) > 0)
    total_outstanding = sum(max(0, int(s.total_fee or 0) - pos_paid.get(s.id, 0)) for s in students)
    active_classes    = sum(1 for c in classes if _enum_val(c.trang_thai) in ("enrolling", "in_progress"))

    # ── Per-branch aggregates ──────────────────────────────────────────────────
    branch_revenue: dict        = defaultdict(int)
    branch_students: dict       = defaultdict(int)
    branch_outstanding: dict    = defaultdict(int)
    branch_students_month: dict = defaultdict(int)

    for p in payments:
        if p.so_tien and int(p.so_tien) > 0 and p.branch_id:
            branch_revenue[p.branch_id] += int(p.so_tien)

    cur_month_start = now_dt.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    for s in students:
        if not s.branch_id:
            continue
        branch_students[s.branch_id] += 1
        bal = max(0, int(s.total_fee or 0) - pos_paid.get(s.id, 0))
        if bal > 0:
            branch_outstanding[s.branch_id] += bal
        ca = s.created_at
        if ca:
            if hasattr(ca, "tzinfo") and ca.tzinfo is None:
                ca = ca.replace(tzinfo=timezone.utc)
            if ca >= cur_month_start:
                branch_students_month[s.branch_id] += 1

    # ── Build workbook ─────────────────────────────────────────────────────────
    wb = Workbook()
    cur_m, cur_y  = now_dt.month, now_dt.year
    prev_m, prev_y = (cur_m - 1, cur_y) if cur_m > 1 else (12, cur_y - 1)

    wb.active.title = "Tổng quan"
    rx.summary_sheet(
        wb.active, now_dt, branches, students, classes,
        branch_map, branch_revenue, branch_students,
        branch_students_month, branch_outstanding,
        status_counts, total_paid, total_outstanding, active_classes,
    )
    rx.student_sheet(wb.create_sheet("Học viên"), students, branch_map, user_map, pos_paid, class_map)
    rx.payment_sheet(wb.create_sheet("Thanh toán"), payments, student_map, branch_map)
    rx.class_sheet(wb.create_sheet("Lớp học"), classes, branch_map)
    ctv_ids = set(ctv_map.keys())
    rx.ctv_students_sheet(wb.create_sheet(f"CTV tháng {cur_m}-{cur_y}"),
                          cur_m, cur_y, students, branch_map, user_map, ctv_ids)
    rx.ctv_students_sheet(wb.create_sheet(f"CTV tháng {prev_m}-{prev_y}"),
                          prev_m, prev_y, students, branch_map, user_map, ctv_ids)

    buf = BytesIO()
    wb.save(buf)
    return _xlsx_resp(buf, f"baocao-{now_dt.strftime('%Y%m%d')}.xlsx")


# ─── CTV competition (xlsx + json) ───────────────────────────────────────────
def _resolve_month_year(month, year) -> tuple[int, int]:
    """Resolve the effective (m, y), defaulting to the current calendar month.

    Cheap (no DB): lets callers build a cache key from the RESOLVED m,y before
    deciding hit/miss, so the default-month case caches under its real key.
    """
    now_dt = datetime.now(timezone.utc)
    return (month or now_dt.month, year or now_dt.year)


async def _load_ctv_ctx(db, month, year):
    """Shared loader for the CTV-competition endpoints.

    Returns (m, y, students, branch_map, user_map, ctv_ids). Defaults month/year
    to the current calendar month. Mirrors data_xlsx's load + lookup pattern.
    """
    from app.models.branch import Branch

    m, y = _resolve_month_year(month, year)

    students = (await db.execute(
        select(Student).where(Student.deleted_at.is_(None)).order_by(Student.created_at.desc())
    )).scalars().all()
    branches  = (await db.execute(select(Branch))).scalars().all()
    all_users = (await db.execute(select(User))).scalars().all()

    branch_map = {b.id: b.ten_chi_nhanh for b in branches}
    user_map   = {u.id: u for u in all_users}
    ctv_ids    = {u.id for u in all_users
                  if u.role == RoleName.collaborator and u.is_active and not u.deleted_at}
    return m, y, students, branch_map, user_map, ctv_ids


@router.get("/ctv-competition.xlsx")
async def ctv_competition_xlsx(current_user: CurrentUser, db: DB,
                               month: int | None = None, year: int | None = None):
    if current_user.role not in (RoleName.admin, RoleName.staff):
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="staff_only")

    m, y, students, branch_map, user_map, ctv_ids = await _load_ctv_ctx(db, month, year)

    wb = Workbook()
    wb.active.title = rx._safe_title(f"CTV tháng {m}-{y}")
    rx.ctv_students_sheet(wb.active, m, y, students, branch_map, user_map, ctv_ids)
    rx.ctv_competition_sheet(wb.create_sheet(rx._safe_title("Bảng vàng CTV")),
                             m, y, students, branch_map, user_map, ctv_ids)
    rx.ctv_podium_sheet(wb.create_sheet(rx._safe_title("Bảng vàng CTV đẹp")),
                        m, y, students, branch_map, user_map, ctv_ids)

    buf = BytesIO()
    wb.save(buf)
    return _xlsx_resp(buf, f"thi-dua-ctv-{m:02d}-{y}.xlsx")


@router.get("/ctv-competition.json")
async def ctv_competition_json(current_user: CurrentUser, db: DB,
                               month: int | None = None, year: int | None = None,
                               fresh: bool = False):
    if current_user.role not in (RoleName.admin, RoleName.staff):
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="staff_only")

    m, y = _resolve_month_year(month, year)
    key = CacheKeys.CTV_COMPETITION.format(year=y, month=m)

    if not fresh:
        try:
            hit = await cache.get(key)
        except Exception:
            hit = None
        if hit is not None:
            return {**hit, "cached": True}

    _, _, students, branch_map, user_map, ctv_ids = await _load_ctv_ctx(db, m, y)
    rows, ctv_count, ctv_rank = rx.compute_ctv_ranking(students, ctv_ids, m, y)

    ordered = [uid for uid, _ in sorted(ctv_rank.items(), key=lambda x: x[1])]

    def _entry(uid):
        staff = user_map.get(uid)
        return {
            "rank": ctv_rank[uid],
            "name": (staff.full_name or staff.email) if staff else "",
            "branch": rx._ctv_branch(uid, rows, user_map, branch_map),
            "count": ctv_count.get(uid, 0),
        }

    payload = {
        "month": m, "year": y, "monthLabel": f"Tháng {m}/{y}",
        "podium": [_entry(uid) for uid in ordered[:3]],
        "rest":   [_entry(uid) for uid in ordered[3:]],
        "totalCtv": len(ordered),
        "totalProfiles": sum(ctv_count.values()),
    }
    try:
        await cache.set(key, payload, ttl=300)
    except Exception:
        pass
    return {**payload, "cached": False}
