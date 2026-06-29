"""Excel workbook sheet builders — called exclusively by reports.py:data_xlsx."""

from __future__ import annotations

from calendar import monthrange
from collections import defaultdict
from datetime import date, datetime, timezone

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# ── Style constants ────────────────────────────────────────────────────────────
_FILL_HDR   = PatternFill("solid", fgColor="0D3B52")
_FILL_SECT  = PatternFill("solid", fgColor="1A5C7A")
_FILL_ALT   = PatternFill("solid", fgColor="F0F7FB")
_FILL_AMBER = PatternFill("solid", fgColor="FFB020")
_FILL_CYAN  = PatternFill("solid", fgColor="00E5FF")
_FILL_LIME  = PatternFill("solid", fgColor="B6FF3C")
_FILL_PINK  = PatternFill("solid", fgColor="FF3D8A")
# Darker pedestal variants (the colored "step" under each pillar).
_FILL_AMBER_D = PatternFill("solid", fgColor="C8860B")
_FILL_CYAN_D  = PatternFill("solid", fgColor="0090A8")
_FILL_LIME_D  = PatternFill("solid", fgColor="7FB81E")
# Deep night backdrop for the podium board (gamey arena feel).
_FILL_NIGHT   = PatternFill("solid", fgColor="0A2533")

_FONT_HDR   = Font(bold=True, color="FFFFFF", size=10)
_FONT_TITLE = Font(bold=True, color="FFFFFF", size=13)
_FONT_SECT  = Font(bold=True, color="FFFFFF", size=10)

# ── Vietnamese display maps ────────────────────────────────────────────────────
_GENDER_VN = {"male": "Nam", "female": "Nữ", "other": "Khác"}
_STATUS_VN = {
    "pending": "Chờ xử lý", "active": "Đang học", "suspended": "Tạm nghỉ",
    "completed": "Đã tốt nghiệp", "dropped": "Nghỉ học",
}
_CLASS_STATUS_VN = {
    "upcoming": "Sắp khai giảng", "enrolling": "Đang tuyển sinh",
    "in_progress": "Đang diễn ra", "completed": "Đã kết thúc", "cancelled": "Đã hủy",
}
_METHOD_VN = {
    "cash": "Tiền mặt", "bank_transfer": "Chuyển khoản",
    "momo": "MoMo", "zalopay": "ZaloPay",
}
_KIND_VN = {"tuition": "Học phí", "rental": "Thuê xe"}


# ── Small formatters ───────────────────────────────────────────────────────────
def _ev(v) -> str:
    return str(v.value if hasattr(v, "value") else (v or ""))

def _vn_date(dt) -> str:
    if dt is None:
        return ""
    return dt.strftime("%d/%m/%Y") if hasattr(dt, "strftime") else str(dt)

def _vn_label(mapping: dict, val) -> str:
    return mapping.get(_ev(val), _ev(val))

def _pay_status_vn(paid: int, total_fee) -> str:
    committed = int(total_fee or 0)
    if committed == 0:    return "Chưa xác định"
    if paid >= committed: return "Đã nộp đủ"
    if paid > 0:          return "Còn nợ"
    return "Chưa nộp"

def _reconstruct_qr(s) -> str:
    """Rebuild an approximate CCCD-QR string from stored fields (legacy students),
    using the OLD on-CCCD address (dia_chi_cccd)."""
    if not s.cccd_number:
        return ""
    dob = s.ngay_sinh.strftime("%d%m%Y") if s.ngay_sinh else ""
    iss = s.cccd_issued_date.strftime("%d%m%Y") if s.cccd_issued_date else ""
    addr = getattr(s, "dia_chi_cccd", None) or s.dia_chi or ""
    return "|".join([s.cccd_number, "", s.ten_hoc_vien or "", dob, _GENDER_VN.get(_ev(s.gioi_tinh), ""), addr, iss])


# ── Low-level sheet primitives ─────────────────────────────────────────────────
def _xhdr(ws, headers: list[str]) -> None:
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        c = ws.cell(1, col)
        c.fill = _FILL_HDR
        c.font = _FONT_HDR
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def _autofit(ws, max_w: int = 40) -> None:
    for col in ws.columns:
        w = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(w + 4, max_w)

_SHEET_TITLE_BAD = str.maketrans({c: "-" for c in r"\/?*[]:"})

def _safe_title(title: str) -> str:
    """Strip characters Excel rejects in sheet names (max 31 chars)."""
    return title.translate(_SHEET_TITLE_BAD)[:31]

def _sect_hdr(ws, label: str) -> int:
    ws.append([label])
    r = ws.max_row
    ws.cell(r, 1).fill = _FILL_SECT
    ws.cell(r, 1).font = _FONT_SECT
    ws.cell(r, 1).alignment = Alignment(vertical="center")
    ws.row_dimensions[r].height = 22
    return r


# ── Sheet 1: Tổng quan (summary) ──────────────────────────────────────────────
def summary_sheet(ws, now_dt, branches, students, classes,
                  branch_map, branch_revenue, branch_students,
                  branch_students_month, branch_outstanding,
                  status_counts, total_paid, total_outstanding, active_classes) -> None:
    # Title
    ws.append(["BÁO CÁO TỔNG QUAN — MOTO GIA THỊNH"])
    ws.cell(1, 1).fill = _FILL_HDR
    ws.cell(1, 1).font = _FONT_TITLE
    ws.cell(1, 1).alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 30

    ws.append([f"Ngày xuất: {now_dt.strftime('%d/%m/%Y %H:%M')}"])
    ws.cell(2, 1).font = Font(italic=True, color="555555", size=9)
    ws.append([])

    # KPI section
    _sect_hdr(ws, "CHỈ SỐ CHUNG")
    ws.append(["CHỈ SỐ", "GIÁ TRỊ"])
    hdr_r = ws.max_row
    for col in [1, 2]:
        ws.cell(hdr_r, col).fill = _FILL_HDR
        ws.cell(hdr_r, col).font = _FONT_HDR
        ws.cell(hdr_r, col).alignment = Alignment(horizontal="center")

    kpis = [
        ("Tổng học viên", len(students)),
        ("Đang học", status_counts.get("active", 0)),
        ("Đã tốt nghiệp", status_counts.get("completed", 0)),
        ("Chờ xử lý", status_counts.get("pending", 0)),
        ("Tạm nghỉ & Nghỉ học", status_counts.get("suspended", 0) + status_counts.get("dropped", 0)),
        ("Tổng doanh thu đã thu (đ)", total_paid),
        ("Tổng còn nợ (đ)", total_outstanding),
        ("Số chi nhánh", len(branches)),
        ("Số lớp đang hoạt động", active_classes),
    ]
    for i, (label, val) in enumerate(kpis):
        ws.append([label, val])
        r = ws.max_row
        if i % 2 == 1:
            for col in [1, 2]:
                ws.cell(r, col).fill = _FILL_ALT
        if "(đ)" in label:
            ws.cell(r, 2).number_format = "#,##0"

    ws.append([])

    # Branch performance section
    _sect_hdr(ws, "HIỆU SUẤT THEO CHI NHÁNH")
    b_hdrs = ["CHI NHÁNH", "TỔNG HV", "HV THÁNG NÀY", "DOANH THU ĐÃ THU (đ)", "CÒN NỢ (đ)"]
    ws.append(b_hdrs)
    hdr_r2 = ws.max_row
    for col in range(1, len(b_hdrs) + 1):
        ws.cell(hdr_r2, col).fill = _FILL_HDR
        ws.cell(hdr_r2, col).font = _FONT_HDR
        ws.cell(hdr_r2, col).alignment = Alignment(horizontal="center")

    for i, b in enumerate(branches):
        ws.append([
            b.ten_chi_nhanh,
            branch_students.get(b.id, 0),
            branch_students_month.get(b.id, 0),
            branch_revenue.get(b.id, 0),
            branch_outstanding.get(b.id, 0),
        ])
        r = ws.max_row
        if i % 2 == 1:
            for col in range(1, 6):
                ws.cell(r, col).fill = _FILL_ALT
        ws.cell(r, 4).number_format = "#,##0"
        ws.cell(r, 5).number_format = "#,##0"

    for letter, w in [("A", 34), ("B", 13), ("C", 16), ("D", 26), ("E", 18)]:
        ws.column_dimensions[letter].width = w


# ── Sheet 2: Học viên (enriched) ──────────────────────────────────────────────
def student_sheet(ws, students, branch_map, user_map, pos_paid, class_map) -> None:
    _xhdr(ws, [
        "MÃ HV", "HỌ TÊN", "NGÀY SINH", "GIỚI TÍNH", "SỐ CCCD",
        "Nơi thường trú - trên CCCD", "Nơi thường trú - địa chỉ mới", "Mã QR CCCD",
        "ĐIỆN THOẠI", "LOẠI BẰNG", "TRẠNG THÁI",
        "NGÀY ĐĂNG KÝ", "CHI NHÁNH", "LỚP HỌC",
        "ĐÃ THANH TOÁN (đ)", "CÒN LẠI (đ)", "TRẠNG THÁI THANH TOÁN",
        "NGƯỜI TẠO", "GHI CHÚ",
    ])
    ws.freeze_panes = "A2"

    for i, s in enumerate(students):
        paid      = pos_paid.get(s.id, 0)
        balance   = max(0, int(s.total_fee or 0) - paid)
        staff     = user_map.get(s.responsible_staff_id)
        staff_name = (staff.full_name or staff.email) if staff else ""

        ws.append([
            s.ma_hoc_vien, s.ten_hoc_vien, _vn_date(s.ngay_sinh),
            _vn_label(_GENDER_VN, s.gioi_tinh), s.cccd_number or "",
            getattr(s, "dia_chi_cccd", None) or "",                 # F: OLD address (on CCCD)
            s.dia_chi or "",                                        # G: NEW address (diachi-converted)
            getattr(s, "cccd_qr_raw", None) or _reconstruct_qr(s),  # H: raw QR (rebuilt for legacy)
            s.so_dien_thoai, _ev(s.loai_bang_lai),
            _vn_label(_STATUS_VN, s.trang_thai),
            _vn_date(s.ngay_dang_ky or s.created_at),
            branch_map.get(s.branch_id, ""),
            class_map.get(s.id, ""),                           # N: Lớp học
            paid, balance, _pay_status_vn(paid, s.total_fee),
            staff_name, s.ghi_chu or "",
        ])
        r = ws.max_row
        if i % 2 == 1:
            for col in range(1, 20):
                ws.cell(r, col).fill = _FILL_ALT
        ws.cell(r, 15).number_format = "#,##0"
        ws.cell(r, 16).number_format = "#,##0"

    _autofit(ws)


# ── Sheet 3: Thanh toán (fixed) ───────────────────────────────────────────────
def payment_sheet(ws, payments, student_map, branch_map) -> None:
    _xhdr(ws, [
        "MÃ GD", "MÃ HV", "TÊN HỌC VIÊN", "CHI NHÁNH",
        "SỐ TIỀN (đ)", "PHƯƠNG THỨC", "LOẠI",
        "NGÀY THU", "SỐ BIÊN LAI", "GHI CHÚ",
    ])
    ws.freeze_panes = "A2"

    for i, p in enumerate(payments):
        s = student_map.get(p.student_id)
        ws.append([
            p.ma_giao_dich,
            s.ma_hoc_vien if s else "",
            s.ten_hoc_vien if s else "",
            branch_map.get(p.branch_id, ""),
            int(p.so_tien) if p.so_tien else 0,
            _vn_label(_METHOD_VN, p.phuong_thuc),
            _vn_label(_KIND_VN, p.kind or "tuition"),
            _vn_date(p.collected_at),
            p.so_bien_lai_id or p.so_bien_lai or "",
            p.ghi_chu or "",
        ])
        r = ws.max_row
        if i % 2 == 1:
            for col in range(1, 11):
                ws.cell(r, col).fill = _FILL_ALT
        ws.cell(r, 5).number_format = "#,##0"

    _autofit(ws)


# ── Sheet 4: Lớp học ──────────────────────────────────────────────────────────
def class_sheet(ws, classes, branch_map) -> None:
    _xhdr(ws, [
        "MÃ LỚP", "TÊN LỚP", "CHI NHÁNH",
        "NGÀY KHAI GIẢNG", "NGÀY KẾT THÚC",
        "SỨC CHỨA", "ĐÃ ĐĂNG KÝ", "TRẠNG THÁI", "GHI CHÚ",
    ])
    ws.freeze_panes = "A2"

    for i, cls in enumerate(classes):
        ws.append([
            cls.ma_lop, cls.ten_lop,
            branch_map.get(cls.branch_id, ""),
            _vn_date(cls.ngay_khai_giang),
            _vn_date(cls.ngay_ket_thuc),
            cls.so_luong_toi_da, cls.so_luong_hien_tai,
            _vn_label(_CLASS_STATUS_VN, cls.trang_thai),
            cls.ghi_chu or "",
        ])
        if i % 2 == 1:
            for col in range(1, 10):
                ws.cell(ws.max_row, col).fill = _FILL_ALT

    _autofit(ws)


# ── Shared CTV ranking (single source of truth) ──────────────────────────────
def _reg(s):
    """A student's registration date (ngay_dang_ky, else created_at's date)."""
    if s.ngay_dang_ky:
        return s.ngay_dang_ky
    ca = s.created_at
    if ca and hasattr(ca, "date"):
        return ca.date()
    return None


def compute_ctv_ranking(students, ctv_ids: set, month: int, year: int):
    """Rank CTVs by number of students registered in the given calendar month.

    Returns (rows, ctv_count, ctv_rank).
    rows: list of month-scoped students whose responsible_staff_id is a ranked CTV,
          sorted by (rank, reg date asc).
    ctv_count: dict uid -> number of students that CTV registered this month.
    ctv_rank: dict uid -> 1-based rank (1 = most students).
    """
    _, last_day = monthrange(year, month)
    m_start = date(year, month, 1)
    m_end   = date(year, month, last_day)

    month_students = [s for s in students if (r := _reg(s)) and m_start <= r <= m_end]

    ctv_count: dict = {}
    for s in month_students:
        uid = s.responsible_staff_id
        if uid and uid in ctv_ids:
            ctv_count[uid] = ctv_count.get(uid, 0) + 1

    ranked_ctv = sorted(ctv_count.items(), key=lambda x: x[1], reverse=True)
    ctv_rank = {uid: rank for rank, (uid, _) in enumerate(ranked_ctv, 1)}

    rows = [s for s in month_students if s.responsible_staff_id in ctv_rank]
    rows.sort(key=lambda s: (ctv_rank[s.responsible_staff_id], _reg(s) or date.min))

    return rows, ctv_count, ctv_rank


# ── Sheets 5 & 6: CTV student details (month-scoped) ─────────────────────────
def ctv_students_sheet(ws, month: int, year: int,
                       students, branch_map, user_map,
                       ctv_ids: set) -> None:
    """Month-scoped CTV performance sheet.

    Shows only CTV-sourced students registered in the given calendar month.
    Columns: MÃ HV · HỌ TÊN · LOẠI BẰNG · NGÀY ĐĂNG KÝ · CHI NHÁNH ·
             NGƯỜI TẠO · SỐ HỒ SƠ · TOP
    Pre-sorted: TOP ASC (rank 1 = most profiles), then NGÀY ĐĂNG KÝ ASC.
    Row fills: rank 1 = amber #FFB020, rank 2 = lime #B6FF3C, rank 3 = cyan #00E5FF,
               rank 4+ = alternating alt/white.
    """
    rows, ctv_count, ctv_rank = compute_ctv_ranking(students, ctv_ids, month, year)

    _xhdr(ws, [
        "MÃ HV", "HỌ TÊN", "LOẠI BẰNG", "NGÀY ĐĂNG KÝ",
        "CHI NHÁNH", "NGƯỜI TẠO", "SỐ HỒ SƠ", "TOP",
    ])
    ws.freeze_panes = "A2"

    for i, s in enumerate(rows):
        rank  = ctv_rank[s.responsible_staff_id]
        staff = user_map.get(s.responsible_staff_id)
        ws.append([
            s.ma_hoc_vien, s.ten_hoc_vien,
            _ev(s.loai_bang_lai),
            _vn_date(s.ngay_dang_ky or s.created_at),
            branch_map.get(s.branch_id, ""),
            (staff.full_name or staff.email) if staff else "",
            ctv_count[s.responsible_staff_id],
            f"TOP {rank}",
        ])
        fill = (
            _FILL_AMBER if rank == 1
            else _FILL_LIME if rank == 2
            else _FILL_CYAN if rank == 3
            else (_FILL_ALT if i % 2 == 1 else None)
        )
        if fill:
            r = ws.max_row
            for col in range(1, 9):
                ws.cell(r, col).fill = fill

    _autofit(ws)


# ── CTV competition banner (top-3 podium + ranked list) ──────────────────────
def _ctv_branch(uid, rows, user_map, branch_map) -> str:
    """Pick a representative branch name for a CTV.

    Uses the branch of that CTV's first (highest-rank, earliest) student row —
    rows are already sorted by (rank, reg date asc). Falls back to the CTV
    user's own branch_id, else "".
    """
    for s in rows:
        if s.responsible_staff_id == uid:
            return branch_map.get(s.branch_id, "")
    staff = user_map.get(uid)
    if staff is not None:
        return branch_map.get(getattr(staff, "branch_id", None), "")
    return ""


def ctv_competition_sheet(ws, month: int, year: int,
                          students, branch_map, user_map,
                          ctv_ids: set) -> None:
    """Trophy-banner report celebrating the top-3 CTVs of the month.

    Shares the single source of truth (compute_ctv_ranking). Renders a tall
    merged trophy banner, a totals subtitle, then a bold bordered podium block
    (rank 1 amber 🥇👑 / rank 2 lime 🥈 / rank 3 cyan 🥉) with name, branch and
    profile count — the champion row reads strongest (thick box, tallest row).
    A "CÁC CTV KHÁC" ranked list follows for rank 4+. Gracefully handles fewer
    than 3 CTVs with a "Chưa có" placeholder. Laid out rank-ascending (1,2,3):
    in a spreadsheet a clean bordered block with the champion highlighted reads
    clearer than a visual 2-1-3 podium.
    """
    rows, ctv_count, ctv_rank = compute_ctv_ranking(students, ctv_ids, month, year)

    # uid ordered by rank (1..N)
    ordered = [uid for uid, _ in sorted(ctv_rank.items(), key=lambda x: x[1])]

    def _name(uid) -> str:
        staff = user_map.get(uid)
        return (staff.full_name or staff.email) if staff else ""

    NCOL = 4  # TOP · NGƯỜI TẠO · CHI NHÁNH · SỐ HỒ SƠ
    total_ctv = len(ordered)
    total_profiles = sum(ctv_count.values())

    # Border sides for the podium block
    thin   = Side(style="thin",   color="0D3B52")
    medium = Side(style="medium", color="0D3B52")
    thick  = Side(style="thick",  color="B8860B")  # dark amber, frames the champion

    # Trophy banner title (merged across all columns, tall)
    ws.append(["🏆  BẢNG VÀNG CTV  🏆"])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=NCOL)
    ws.cell(1, 1).fill = _FILL_HDR
    ws.cell(1, 1).font = Font(bold=True, color="FFFFFF", size=18)
    ws.cell(1, 1).alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 46

    # Month subtitle
    ws.append([f"Tháng {month} / {year}"])
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=NCOL)
    ws.cell(2, 1).fill = _FILL_SECT
    ws.cell(2, 1).font = Font(bold=True, color="FFFFFF", size=12)
    ws.cell(2, 1).alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 24

    # Totals subtitle (mirrors the dialog footer)
    ws.append([f"{total_ctv} CTV · {total_profiles} hồ sơ"])
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=NCOL)
    ws.cell(3, 1).font = Font(italic=True, color="555555", size=10)
    ws.cell(3, 1).alignment = Alignment(horizontal="center", vertical="center")
    ws.append([])

    # Podium header
    ws.append(["TOP", "NGƯỜI TẠO", "CHI NHÁNH", "SỐ HỒ SƠ"])
    hdr_r = ws.max_row
    for col in range(1, NCOL + 1):
        ws.cell(hdr_r, col).fill = _FILL_HDR
        ws.cell(hdr_r, col).font = _FONT_HDR
        ws.cell(hdr_r, col).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(hdr_r, col).border = Border(left=medium, right=medium, top=medium, bottom=thin)
    ws.row_dimensions[hdr_r].height = 20

    podium = {
        1: (_FILL_AMBER, "🥇  TOP 1",    13, 36),
        2: (_FILL_LIME,  "🥈  TOP 2",    12, 30),
        3: (_FILL_CYAN,  "🥉  TOP 3",    12, 30),
    }
    for rank in (1, 2, 3):
        fill, label, fsize, height = podium[rank]
        if rank <= len(ordered):
            uid = ordered[rank - 1]
            ws.append([label, _name(uid),
                       _ctv_branch(uid, rows, user_map, branch_map),
                       ctv_count.get(uid, 0)])
        else:
            ws.append([label, "Chưa có", "", 0])
        r = ws.max_row
        bottom = medium if rank == 3 else thin
        for col in range(1, NCOL + 1):
            cell = ws.cell(r, col)
            cell.fill = fill
            cell.font = Font(bold=True, size=fsize, color="0D3B52")
            cell.alignment = Alignment(
                horizontal="center" if col in (1, 4) else "left", vertical="center")
            # Champion (rank 1) gets a thick amber frame; others a medium box.
            if rank == 1:
                cell.border = Border(
                    left=thick if col == 1 else medium,
                    right=thick if col == NCOL else medium,
                    top=thick, bottom=thick)
            else:
                cell.border = Border(
                    left=medium, right=medium,
                    top=thin, bottom=bottom)
        ws.row_dimensions[r].height = height

    ws.append([])

    # Remaining CTVs (rank 4+)
    _sect_hdr(ws, "CÁC CTV KHÁC")
    ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=NCOL)
    ws.append(["TOP", "NGƯỜI TẠO", "CHI NHÁNH", "SỐ HỒ SƠ"])
    hdr_r2 = ws.max_row
    for col in range(1, NCOL + 1):
        ws.cell(hdr_r2, col).fill = _FILL_HDR
        ws.cell(hdr_r2, col).font = _FONT_HDR
        ws.cell(hdr_r2, col).alignment = Alignment(horizontal="center")

    rest = ordered[3:]
    if rest:
        for i, uid in enumerate(rest):
            rank = ctv_rank[uid]
            ws.append([f"TOP {rank}", _name(uid),
                       _ctv_branch(uid, rows, user_map, branch_map),
                       ctv_count.get(uid, 0)])
            r = ws.max_row
            ws.cell(r, 1).alignment = Alignment(horizontal="center")
            if i % 2 == 1:
                for col in range(1, NCOL + 1):
                    ws.cell(r, col).fill = _FILL_ALT
    else:
        ws.append(["", "Chưa có", "", ""])

    for letter, w in [("A", 14), ("B", 34), ("C", 26), ("D", 12)]:
        ws.column_dimensions[letter].width = w


# ── CTV podium board (gamey visual 2-1-3 podium) ─────────────────────────────
def ctv_podium_sheet(ws, month: int, year: int,
                     students, branch_map, user_map,
                     ctv_ids: set) -> None:
    """Gamey visual podium board — the spreadsheet twin of the in-app dialog.

    A real 2-1-3 podium: three colored pillars side by side on a deep-night
    arena, champion in the center raised highest (amber), rank 2 left (lime),
    rank 3 right (cyan). Each pillar is a vertical stack of merged rows so the
    profile count reads as a big hero number, the name medium, the branch small.
    A colored pedestal "step" with 🥇/🥈/🥉 sits under each pillar, then a
    "CÁC CTV KHÁC" ranked list for rank 4+. Gracefully handles <3 CTVs.

    Shares compute_ctv_ranking (single source of truth) with the list sheets.
    Grid: 6 columns, each pillar spans 2 (B:C left, D:E center, F:G right);
    col A is a thin night gutter so the board floats.
    """
    rows, ctv_count, ctv_rank = compute_ctv_ranking(students, ctv_ids, month, year)
    ordered = [uid for uid, _ in sorted(ctv_rank.items(), key=lambda x: x[1])]

    def _name(uid) -> str:
        staff = user_map.get(uid)
        return (staff.full_name or staff.email) if staff else ""

    def _entry(rank: int):
        if rank <= len(ordered):
            uid = ordered[rank - 1]
            return _name(uid), _ctv_branch(uid, rows, user_map, branch_map), ctv_count.get(uid, 0)
        return "Chưa có", "", 0

    NCOL = 8  # A gutter + 3 pillars × 2 cols + H gutter (mirror of A)
    total_ctv = len(ordered)
    total_profiles = sum(ctv_count.values())

    night_side = Side(style="thin", color="0A2533")
    gold_side  = Side(style="thick", color="FFD24A")

    def _fill_row(r: int, fill):
        for col in range(1, NCOL + 1):
            ws.cell(r, col).fill = fill

    def _merge(r1, c1, r2, c2):
        ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)

    # ── Banner ───────────────────────────────────────────────────────────────
    ws.append(["🏆   B Ả N G   V À N G   C T V   🏆"])
    _merge(1, 1, 1, NCOL)
    c = ws.cell(1, 1)
    c.fill = _FILL_AMBER; c.font = Font(bold=True, color="0A2533", size=20)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 52

    ws.append([f"THÁNG {month} / {year}   ·   {total_ctv} CTV  ·  {total_profiles} hồ sơ"])
    _merge(2, 1, 2, NCOL)
    c = ws.cell(2, 1)
    c.fill = _FILL_HDR; c.font = Font(bold=True, color="FFFFFF", size=12)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 26

    # Night gap row under the banner.
    ws.append([]); _fill_row(ws.max_row, _FILL_NIGHT); ws.row_dimensions[ws.max_row].height = 15

    # ── Podium pillars ─────────────────────────────────────────────────────────
    # The board sits on a fixed grid of rows; the champion column starts two
    # rows higher than the flanks so it physically stands tallest.
    base = ws.max_row + 1                      # first board row
    PILLAR_ROWS = 7                            # crown/medal · name · branch · count · count-label · pad · pad
    TOTAL_ROWS = PILLAR_ROWS + 2               # + champion's 2-row head start
    board_rows = list(range(base, base + TOTAL_ROWS))
    for r in board_rows:
        ws.append([])
        _fill_row(r, _FILL_NIGHT)
        ws.row_dimensions[r].height = 22

    # Each pillar: (rank, first_col, fill, pedestal_fill, medal, head_start)
    pillars = [
        (2, 2, _FILL_LIME,  _FILL_LIME_D,  "🥈", 2),   # left  (B:C)
        (1, 4, _FILL_AMBER, _FILL_AMBER_D, "🥇", 0),   # center(D:E) — tallest
        (3, 6, _FILL_CYAN,  _FILL_CYAN_D,  "🥉", 2),   # right (F:G)
    ]

    for rank, c1, fill, ped_fill, medal, head in pillars:
        c2 = c1 + 1
        name, branch, count = _entry(rank)
        top = base + head                      # champion starts higher
        body_side = gold_side if rank == 1 else night_side

        # Layout rows inside this pillar
        r_medal = top
        r_name  = top + 1
        r_brnch = top + 2
        r_count = top + 3
        r_clbl  = top + 4
        r_ped   = base + TOTAL_ROWS - 1        # pedestals align at the very bottom

        # Crown + medal head
        _merge(r_medal, c1, r_medal, c2)
        cell = ws.cell(r_medal, c1)
        cell.value = ("👑 " + medal) if rank == 1 else medal
        cell.fill = fill
        cell.font = Font(bold=True, size=20 if rank == 1 else 16, color="0A2533")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[r_medal].height = 30 if rank == 1 else 24

        # Name
        _merge(r_name, c1, r_name, c2)
        cell = ws.cell(r_name, c1)
        cell.value = name
        cell.fill = fill
        cell.font = Font(bold=True, size=12 if rank == 1 else 11, color="0A2533")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Branch
        _merge(r_brnch, c1, r_brnch, c2)
        cell = ws.cell(r_brnch, c1)
        cell.value = branch
        cell.fill = fill
        cell.font = Font(italic=True, size=9, color="0A2533")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Hero count
        _merge(r_count, c1, r_count, c2)
        cell = ws.cell(r_count, c1)
        cell.value = count
        cell.fill = fill
        cell.font = Font(bold=True, size=30 if rank == 1 else 24, color="0A2533")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[r_count].height = 40 if rank == 1 else 34

        # "hồ sơ" caption
        _merge(r_clbl, c1, r_clbl, c2)
        cell = ws.cell(r_clbl, c1)
        cell.value = "HỒ SƠ"
        cell.fill = fill
        cell.font = Font(bold=True, size=9, color="0A2533")
        cell.alignment = Alignment(horizontal="center", vertical="center")

        # Fill body (between caption and pedestal) with the pillar color so it
        # reads as one solid column.
        for r in range(r_clbl + 1, r_ped):
            for col in (c1, c2):
                ws.cell(r, col).fill = fill

        # Pedestal step — darker shade, big rank label.
        _merge(r_ped, c1, r_ped, c2)
        cell = ws.cell(r_ped, c1)
        cell.value = f"TOP {rank}"
        cell.fill = ped_fill
        cell.font = Font(bold=True, size=14 if rank == 1 else 12, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[r_ped].height = 30 if rank == 1 else 24

        # Champion gets a thick gold frame around its whole pillar.
        if rank == 1:
            for r in range(r_medal, r_ped + 1):
                ws.cell(r, c1).border = Border(
                    left=gold_side,
                    right=Side(style="thin", color="C8860B"),
                    top=gold_side if r == r_medal else None,
                    bottom=gold_side if r == r_ped else None)
                ws.cell(r, c2).border = Border(
                    left=Side(style="thin", color="C8860B"),
                    right=gold_side,
                    top=gold_side if r == r_medal else None,
                    bottom=gold_side if r == r_ped else None)

    # Night floor strip below the podium.
    ws.append([]); _fill_row(ws.max_row, _FILL_NIGHT); ws.row_dimensions[ws.max_row].height = 15

    # ── CÁC CTV KHÁC (rank 4+) ─────────────────────────────────────────────────
    # Same list as sheet 2 (ctv_competition_sheet): TOP · NGƯỜI TẠO · CHI NHÁNH ·
    # SỐ HỒ SƠ, _FILL_HDR header, _FILL_ALT striping on odd rows, only TOP centered.
    # The four fields are mapped onto the podium's 8-col grid two columns each
    # (TOP=A:B, name=C:D, branch=E:F, count=G:H) — columns are shared with the
    # podium so sheet 2's exact widths can't apply, but the structure and styling do.
    ws.append([])
    _sect_hdr(ws, "CÁC CTV KHÁC")
    _merge(ws.max_row, 1, ws.max_row, NCOL)

    ws.append(["TOP", "", "NGƯỜI TẠO", "", "CHI NHÁNH", "", "SỐ HỒ SƠ", ""])
    hdr_r = ws.max_row
    _merge(hdr_r, 1, hdr_r, 2); _merge(hdr_r, 3, hdr_r, 4)
    _merge(hdr_r, 5, hdr_r, 6); _merge(hdr_r, 7, hdr_r, 8)
    for col in range(1, NCOL + 1):
        ws.cell(hdr_r, col).fill = _FILL_HDR
        ws.cell(hdr_r, col).font = _FONT_HDR
        ws.cell(hdr_r, col).alignment = Alignment(horizontal="center")

    rest = ordered[3:]
    if rest:
        for i, uid in enumerate(rest):
            rank = ctv_rank[uid]
            ws.append([f"TOP {rank}", "", _name(uid), "",
                       _ctv_branch(uid, rows, user_map, branch_map), "",
                       ctv_count.get(uid, 0), ""])
            r = ws.max_row
            _merge(r, 1, r, 2); _merge(r, 3, r, 4)
            _merge(r, 5, r, 6); _merge(r, 7, r, 8)
            ws.cell(r, 1).alignment = Alignment(horizontal="center")
            if i % 2 == 1:
                _fill_row(r, _FILL_ALT)
    else:
        ws.append(["", "", "Chưa có", "", "", "", "", ""])
        _merge(ws.max_row, 3, ws.max_row, NCOL)

    # Column widths — gutters (A, H) thin, pillars even. A and H are the
    # night side-margins that frame the board symmetrically.
    for letter, w in [("A", 6), ("B", 16), ("C", 16), ("D", 16),
                      ("E", 16), ("F", 16), ("G", 16), ("H", 6)]:
        ws.column_dimensions[letter].width = w
